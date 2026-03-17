
import logging
from puresnmp import get
import socket
from nmap import PortScanner
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import concurrent.futures
from datetime import datetime

class SNMPManager:
    def __init__(self):
        self.stop_flag = False

    def stop(self):
        self.stop_flag = True

    def run_snmp_scan(self, network, config, queue_obj):
        self.stop_flag = False
        nm = PortScanner()
        if '/' not in str(network):
            network = str(network) + '/24'

        try:
            nm.scan(network, arguments='-sP -n -PE --max-rate 100')
        except Exception as e:
             if queue_obj:
                 queue_obj.put({'type': 'error', 'text': f"Nmap Error: {e}"})
                 queue_obj.put({'type': 'done'})
             return

        hosts = nm.all_hosts()
        if not hosts:
             if queue_obj: queue_obj.put({'type': 'log', 'text': "⚠️ Aucun hôte ne répond au Ping. Tentative de découverte SNMP directe...", 'tag': 'warning'})
             try:
                 import ipaddress
                 net_obj = ipaddress.ip_network(network, strict=False)
                 if net_obj.num_addresses <= 256:
                     hosts = [str(ip) for ip in net_obj.hosts()]
                 else:
                     if queue_obj: 
                         queue_obj.put({'type': 'error', 'text': f"Réseau trop large ({net_obj.num_addresses} IPs) pour une découverte sans Ping."})
                         queue_obj.put({'type': 'done'})
                     return
             except Exception:
                 if queue_obj:
                     queue_obj.put({'type': 'error', 'text': f"Aucun hôte répondant au Ping sur {network}."})
                     queue_obj.put({'type': 'done'})
                 return
        
        excel_filename = f'scan_{network.replace("/", "_")}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scan Results"
        ws.append(['Date-Heure', 'AdresseIP', 'Version', 'AdresseMAC', 'Nom', 'Modèle', 'Description', 'Localisation', 'Contact'])
        
        for cell in ws[1]: cell.font = Font(bold=True)

        communities = config['settings'].get('snmp_communities', ['public', 'TICE'])
        oid_map = config.get('oids', {})
        ip_limit = config['settings'].get('ip_scan_limit_last_octet', 254)

        valid_hosts = []
        for host in hosts:
            if self.stop_flag: break
            try:
                last_octet = int(host.split('.')[-1])
                if last_octet > ip_limit and last_octet != 254:
                    continue
            except: pass
            valid_hosts.append(host)

        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
            future_to_host = {executor.submit(self.process_single_host, host, communities, oid_map): host for host in valid_hosts}
            
            completed_count = 0
            for future in concurrent.futures.as_completed(future_to_host):
                if self.stop_flag: break
                completed_count += 1
                result = future.result()
                if result:
                    results.append(result)
                    if queue_obj:
                        queue_obj.put({
                            'type': 'result',
                            'ip': result[1],
                            'snmp': result[2],
                            'mac': result[3],
                            'name': result[4],
                            'model': result[5],
                            'desc': result[6],
                            'location': result[7],
                            'contact': result[8] if len(result) > 8 else ''
                        })
                
                if queue_obj:
                    progress_pct = int((completed_count / len(valid_hosts)) * 100) if valid_hosts else 100
                    queue_obj.put({'type': 'progress', 'value': progress_pct})

        results.sort(key=lambda x: int(x[1].split('.')[-1]) if x[1] and x[1].count('.')==3 else 0)
        for row in results: ws.append(row)

        try:
            tab = Table(displayName="ScanResults", ref=f"A1:I{len(results)+1}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tab)
        except: pass

        for column in ws.columns:
            try:
                max_len = max([len(str(c.value)) for c in column if c.value] or [0])
                ws.column_dimensions[get_column_letter(column[0].column)].width = max_len + 2
            except: pass

        wb.save(excel_filename)
        
        # Save IP cache for Offline Mode
        try:
            import json
            from server.utils.paths import get_file_path
            cache_ips = [row[1] for row in results if len(row) > 1 and row[1]]
            with open(get_file_path("last_inventory.json"), "w", encoding='utf-8') as f:
                json.dump(cache_ips, f)
        except Exception:
            pass # Non-blocking

        if queue_obj:
            queue_obj.put({'type': 'done', 'file': excel_filename})

    def process_single_host(self, host, communities, oid_map):
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        for community in communities:
            # Try SNMP v2c (most common) then v1 (older devices)
            for snmp_ver in [2, 1]:
                try:
                    info = self.snmp_get_info(host, community, oid_map, version=snmp_ver)
                    if info:
                        ver_lbl = "v2c" if snmp_ver == 2 else "v1"
                        return [now_str, host, ver_lbl, *info]
                except (socket.timeout, Exception):
                    continue
        
        return [now_str, host, 'NO SNMP', '', '', '', '', '', '']

    def snmp_get_info(self, host, community, oid_map, version=2):
        def fetch(oid):
            try:
                # puresnmp.get handles OIDs best without leading dot usually, but let's be safe
                clean_oid = oid.lstrip('.')
                res = get(host, community, clean_oid, timeout=3, retries=1, version=version)
                if res is None: return ""
                
                # Handle bytes (OctetString)
                if isinstance(res, bytes):
                    return res.decode('utf-8', errors='ignore')
                
                # Handle other types (Integer, OID, etc.)
                return str(res)
            except:
                return ""

        def clean(val):
            s = str(val)
            # Remove "b'..." prefix if it's there
            if s.startswith("b'") and s.endswith("'"): s = s[2:-1]
            return "".join(ch for ch in s if (31 < ord(ch) < 127) or ord(ch) > 160 or ch in '\t\n\r')

        # 1. System Descr (Standard OID)
        sys_descr = clean(fetch('1.3.6.1.2.1.1.1.0'))
        if not sys_descr or len(sys_descr) < 2:
             return None
        
        # 2. MAC Address (Some devices return it here)
        raw_mac = fetch('1.3.6.1.2.1.17.1.1.0')
        mac_addr = "N/A"
        if raw_mac:
            # If it's pure bytes and 6 bytes long, it's a MAC
            if isinstance(raw_mac, bytes) and len(raw_mac) == 6:
                mac_addr = ":".join(f"{b:02X}" for b in raw_mac)
            else:
                s_mac = clean(raw_mac)
                if len(s_mac) > 5: mac_addr = s_mac
  
        # 3. Model
        sys_obj_id = fetch('1.3.6.1.2.1.1.2.0')
        model_name = "Inconnu"
        if sys_obj_id:
             s_oid = str(sys_obj_id).lstrip('.')
             for key, label in oid_map.items():
                 if key.lstrip('.') in s_oid:
                     model_name = label
                     break
             if model_name == "Inconnu":
                 model_name = f"OID: {s_oid}"

        # 4. Contact, Name, Location
        contact = clean(fetch('1.3.6.1.2.1.1.4.0'))
        name = clean(fetch('1.3.6.1.2.1.1.5.0'))
        location = clean(fetch('1.3.6.1.2.1.1.6.0'))

        # Heuristic for unknown models
        if "Inconnu" in model_name or "OID:" in model_name:
            if "Cisco" in sys_descr: model_name = "Cisco Device"
            elif "Aruba" in sys_descr: model_name = "Aruba Device"
            elif "HPE" in sys_descr: model_name = "HPE Device"

        return [mac_addr, name, model_name, sys_descr, location, contact]
