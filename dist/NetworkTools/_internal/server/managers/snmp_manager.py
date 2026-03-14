
from pysnmp.hlapi import SnmpEngine, CommunityData, UdpTransportTarget, ContextData, ObjectType, ObjectIdentity, getCmd
from nmap import PortScanner
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import concurrent.futures
from datetime import datetime

class SNMPManager:
    def __init__(self):
        self.stop_flag = False

    def stop(self):
        self.stop_flag = True

    def run_snmp_scan(self, network, config, queue_obj):
        self.stop_flag = False
        nm = PortScanner()
        # Handle CIDR: if user didn't provide /xx, add /24 default
        if '/' not in str(network):
            network = str(network) + '/24'

        try:
            # Ping scan only with conservative rate limiting (Anti-DDOS/IPS protection)
            nm.scan(network, arguments='-sP -n -PE --max-rate 100')
        except Exception as e:
             if queue_obj:
                 queue_obj.put({'type': 'error', 'text': f"Nmap Error: {e}"})
                 queue_obj.put({'type': 'done'}) # Unblock UI
             return

        hosts = nm.all_hosts()
        if not hosts:
             if queue_obj:
                 queue_obj.put({'type': 'error', 'text': f"Aucun hôte répondant au Ping sur {network}."})
                 queue_obj.put({'type': 'done'}) # Unblock UI
             return
        
        excel_filename = f'scan_{network.replace("/", "_")}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scan Results"
        ws.append(['Date-Heure', 'AdresseIP', 'Version', 'AdresseMAC', 'Nom', 'Modèle', 'Description', 'Localisation'])
        
        for cell in ws[1]: cell.font = Font(bold=True)

        communities = config['settings']['snmp_communities']
        oid_map = config['oids']
        ip_limit = config['settings'].get('ip_scan_limit_last_octet', 254)

        # Initialize valid hosts list
        valid_hosts = []
        for host in hosts:
            if self.stop_flag:
                if queue_obj: queue_obj.put({'type': 'log', 'text': f"Scan annulé avant filtre."})
                break
                
            try:
                last_octet = int(host.split('.')[-1])
                if last_octet > ip_limit and last_octet != 254:
                    continue
            except: pass
            valid_hosts.append(host)

        if self.stop_flag:
            if queue_obj: queue_obj.put({'type': 'done', 'message': "Scan SNMP Annulé."})
            return

        # Parallel Execution
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
            future_to_host = {}
            for host in valid_hosts:
                if self.stop_flag: break
                future_to_host[executor.submit(self.process_single_host, host, communities, oid_map)] = host
            
            completed_count = 0
            for future in concurrent.futures.as_completed(future_to_host):
                if self.stop_flag: 
                    # Attempt to cancel remaining not yet started futures
                    for f in future_to_host: f.cancel()
                    break
                    
                completed_count += 1
                result = future.result()
                if result:
                    results.append(result)
                    # New: Send real-time result to UI
                    if queue_obj:
                        queue_obj.put({
                            'type': 'result',
                            'ip': result[1],
                            'snmp': result[2],
                            'mac': result[3],
                            'name': result[4],
                            'model': result[5],
                            'desc': result[6],
                            'location': result[7]
                        })
                
                # Update progress
                if queue_obj:
                    progress_pct = int((completed_count / len(valid_hosts)) * 100) if valid_hosts else 100
                    queue_obj.put({'type': 'progress', 'value': progress_pct})

        # Write results to Excel
        results.sort(key=lambda x: int(x[1].split('.')[-1]) if x[1] and x[1].count('.')==3 else 0)

        for row in results:
            ws.append(row)

        # Format Table
        try:
            tab = Table(displayName="ScanResults", ref=f"A1:H{len(results)+1}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tab)
        except ValueError: pass # Empty table or error

        # Auto-width
        for column in ws.columns:
            try:
                max_len = max([len(str(c.value)) for c in column if c.value] or [0])
                ws.column_dimensions[get_column_letter(column[0].column)].width = max_len + 2
            except: pass

        wb.save(excel_filename)
        
        # Save IP cache for Offline Mode
        try:
            import json
            from server.utils.paths import get_file_path
            cache_ips = [row[1] for row in results if len(row) > 1 and row[1]]
            with open(get_file_path("last_inventory.json"), "w", encoding='utf-8') as f:
                json.dump(cache_ips, f)
        except Exception:
            pass # Non-blocking

        if queue_obj:
            queue_obj.put({'type': 'done', 'file': excel_filename})

    def process_single_host(self, host, communities, oid_map):
        for community in communities:
            # Try both v2 (mpModel=1) and v1 (mpModel=0)
            for version_code, version_lbl in [(1, 'V2'), (0, 'V1')]:
                auth_data = CommunityData(community, mpModel=version_code)
                try:
                    info = self.snmp_get_info(host, auth_data, oid_map)
                    if info:
                        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                        return [now_str, host, version_lbl, *info]
                except Exception:
                    continue
        
        # If no SNMP found
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        return [now_str, host, 'NO SNMP', '', '', '', '', '']

    def snmp_get_info(self, host, auth, oid_map):
        engine = SnmpEngine()
        target = UdpTransportTarget((host, 161), timeout=1, retries=1)
        
        def clean_string(val):
             if not val: return ""
             return "".join(ch for ch in str(val) if (31 < ord(ch) < 127) or ord(ch) > 160 or ch in '\t\n\r')

        def fetch_oid(oid_str):
            try:
                g = getCmd(engine, auth, target, ContextData(), ObjectType(ObjectIdentity(oid_str)))
                errorIndication, errorStatus, errorIndex, varBinds = next(g)
                if errorIndication or errorStatus: return None
                return varBinds[0][1]
            except: return None

        # 1. System Descr
        val = fetch_oid('1.3.6.1.2.1.1.1.0')
        if not val: return None
        sys_descr = clean_string(val.prettyPrint())
        
        # 2. MAC Address
        val_mac = fetch_oid('1.3.6.1.2.1.17.1.1.0')
        mac_addr = "N/A"
        if val_mac:
            try:
                if hasattr(val_mac, 'asNumbers'):
                    nums = val_mac.asNumbers()
                    if nums:
                        hex_str = "".join(["{:02x}".format(x) for x in nums])
                        mac_addr = ":".join(hex_str[i:i+2] for i in range(0, len(hex_str), 2)).upper()
                else:
                    mac_addr = clean_string(val_mac.prettyPrint())
            except:
                mac_addr = clean_string(val_mac.prettyPrint())

        # 3. Model
        val_oid = fetch_oid('1.3.6.1.2.1.1.2.0')
        model_name = "Inconnu"
        clean_oid = ""
        if val_oid:
             clean_oid = str(val_oid.prettyPrint())
             for key, label in oid_map.items():
                 if key in clean_oid:
                     model_name = label
                     break
        
        # 4. Contact, Name, Location
        contact = clean_string(fetch_oid('1.3.6.1.2.1.1.4.0').prettyPrint()) if fetch_oid('1.3.6.1.2.1.1.4.0') else ""
        name = clean_string(fetch_oid('1.3.6.1.2.1.1.5.0').prettyPrint()) if fetch_oid('1.3.6.1.2.1.1.5.0') else ""
        location = clean_string(fetch_oid('1.3.6.1.2.1.1.6.0').prettyPrint()) if fetch_oid('1.3.6.1.2.1.1.6.0') else ""

        if model_name == "Inconnu" and clean_oid:
            model_name = f"OID: {clean_oid}"
            if sys_descr and "Cisco" in sys_descr:
                 model_name = "Cisco Device (Unknown OID)"
            elif sys_descr and "Aruba" in sys_descr:
                 model_name = "Aruba Device (Unknown OID)"
        
        return [mac_addr, name, model_name, sys_descr, location]
